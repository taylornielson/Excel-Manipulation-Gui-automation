import openpyxl, os, calendar, subprocess, pyautogui
from openpyxl.styles import PatternFill, Color

os.chdir('C:\\Users\\nielson\\Desktop')
curMonth = 9

def fillrow(ind, sheet_name):
    for x in range(1, sheet_name.max_column):
        sheet_name.cell(row = ind, column = x).fill = greenfill

gpInfo = {}
keys = []
wb = openpyxl.load_workbook(os.getcwd() + '\\todays_batch.xlsx')
print(os.getcwd())
sheet = wb['Transaction Entry Distributions']
greenfill = PatternFill(fill_type= 'solid', start_color=Color('00CC00'), end_color=Color('00CC00'))
sheet.insert_cols(2)
sheet.cell(row = 1, column = 2).value = "FAC"
for i in range(2, sheet.max_row + 1):
        sheet.cell(row= i, column = 2).value = "=LEFT(A" + str(i) + ",3)"
sheet.insert_cols(9)
sheet.cell(row = 1, column = 9).value = "FAC"
for i in range(2, sheet.max_row + 1):
        sheet.cell(row= i, column = 9).value = "=LEFT(H" + str(i) + ",3)"
sheet.insert_cols(10)
sheet.cell(row = 1, column = 10).value = "MATCH"
for i in range(2, sheet.max_row + 1):
        sheet.cell(row= i, column = 10).value = "=IF(B" + str(i) + "=I" + str(i) +",0,1)" 
for i in range(1, sheet.max_row + 1):
    val = sheet.cell(row=i, column = 5).value.split(',')
    if len(val) != 1:
        if int(val[1]) > curMonth:
                fillrow(i, sheet)
                if "IPY_UNPAID" in sheet.cell(row = i, column = 16).value:
                        keys.append(sheet.cell(row = i, column = 19).value)
                month = str(calendar.month_name[int(val[1])])
                sheet.cell(row = i, column = 16).value = ('IPY_' + month.upper())
for i in range(2, sheet.max_row + 1):
        if sheet.cell(row = i, column = 19).value in keys:
                key = sheet.cell(row = i, column = 19).value
                value = sheet.cell(row=i, column = 16).value
                gpInfo.update({key:value})
            
print (gpInfo)

wb.save('C:\\Users\\nielson\\Desktop\\example_output.xlsx')
pyautogui.PAUSE = 1.0
pyautogui.moveTo(569,1200); pyautogui.click(569, 1039); pyautogui.moveTo(83,70); pyautogui.click(); 
pyautogui.PAUSE=0.5
pyautogui.moveTo(90,130); pyautogui.moveTo(250, 130); pyautogui.click(); pyautogui.moveTo(350,183)
for key in gpInfo:
        # Need to create check for adding new batch
       pyautogui.moveTo(330,250); pyautogui.press("backspace"); pyautogui.typewrite(key); pyautogui.press("tab"); pyautogui.moveTo(650,250); pyautogui.click(650,250); pyautogui.typewrite('IPY'); pyautogui.keyDown('shift'); pyautogui.press('-'); pyautogui.keyUp('shift'); pyautogui.typewrite(value[4:]); pyautogui.press("enter"); pyautogui.press("enter")
